"""
Excel Rapor Modülü
Filtrelenebilir Excel raporları oluşturur
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
import datetime


class ExcelRaporlama:
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def rapor_olustur(self, sikayetler, dosya_adi=None):
        """
        Şikayetlerden Excel raporu oluşturur
        
        Args:
            sikayetler (list): Şikayet listesi (tuple formatında)
            dosya_adi (str): Kaydedilecek dosya adı
            
        Returns:
            str: Oluşturulan dosya yolu
        """
        if not dosya_adi:
            tarih = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            dosya_adi = f"Sikayet_Raporu_{tarih}.xlsx"
        
        self.wb = Workbook()
        
        # 1. Sayfa: Detaylı Liste
        self._detayli_liste_olustur(sikayetler)
        
        # 2. Sayfa: Özet İstatistikler
        self._ozet_istatistikler_olustur(sikayetler)
        
        # 3. Sayfa: Grafikler
        self._grafikler_olustur(sikayetler)
        
        # Kaydet
        self.wb.save(dosya_adi)
        return dosya_adi
    
    def _detayli_liste_olustur(self, sikayetler):
        """Detaylı şikayet listesi sayfası"""
        self.ws = self.wb.active
        self.ws.title = "Şikayet Listesi"
        
        # Başlıklar
        basliklar = [
            "Şikayet No", "Yolcu Adı", "Telefon", "E-posta",
            "Seyahat Tarihi", "Güzergah", "PNR", "Plaka",
            "Şikayet Türü", "Öncelik", "Durum", "Kayıt Tarihi",
            "Şikayet Detayı"
        ]
        
        # Başlık satırını yaz
        for col, baslik in enumerate(basliklar, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = baslik
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Veri satırlarını yaz
        for row_idx, sikayet in enumerate(sikayetler, 2):
            # sikayet tuple formatı: (id, sikayet_no, yolcu_adi, seyahat_tarihi, guzergah, pnr, iletisim, platform, sikayet_detay, kayit_tarihi, durum, telefon, eposta, plaka, sikayet_turu, lokasyon, oncelik, ...)
            
            self.ws.cell(row=row_idx, column=1).value = sikayet[1]  # sikayet_no
            self.ws.cell(row=row_idx, column=2).value = sikayet[2]  # yolcu_adi
            self.ws.cell(row=row_idx, column=3).value = sikayet[11] if len(sikayet) > 11 else ""  # telefon
            self.ws.cell(row=row_idx, column=4).value = sikayet[12] if len(sikayet) > 12 else ""  # eposta
            self.ws.cell(row=row_idx, column=5).value = sikayet[3]  # seyahat_tarihi
            self.ws.cell(row=row_idx, column=6).value = sikayet[4]  # guzergah
            self.ws.cell(row=row_idx, column=7).value = sikayet[5]  # pnr
            self.ws.cell(row=row_idx, column=8).value = sikayet[13] if len(sikayet) > 13 else ""  # plaka
            self.ws.cell(row=row_idx, column=9).value = sikayet[14] if len(sikayet) > 14 else ""  # sikayet_turu
            self.ws.cell(row=row_idx, column=10).value = sikayet[16] if len(sikayet) > 16 else ""  # oncelik
            self.ws.cell(row=row_idx, column=11).value = sikayet[10]  # durum
            self.ws.cell(row=row_idx, column=12).value = sikayet[9]  # kayit_tarihi
            self.ws.cell(row=row_idx, column=13).value = sikayet[8]  # sikayet_detay
            
            # Durum renklendir
            durum_cell = self.ws.cell(row=row_idx, column=11)
            if sikayet[10] == "Yeni":
                durum_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            elif sikayet[10] == "İşlemde":
                durum_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif sikayet[10] == "Çözüldü":
                durum_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            
            # Öncelik renklendir
            oncelik_cell = self.ws.cell(row=row_idx, column=10)
            oncelik = sikayet[16] if len(sikayet) > 16 else ""
            if oncelik == "Acil":
                oncelik_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            elif oncelik == "Yüksek":
                oncelik_cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        column_widths = [15, 20, 15, 25, 15, 20, 12, 12, 20, 12, 12, 18, 50]
        for col, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[get_column_letter(col)].width = width
        
        # Filtre ekle
        self.ws.auto_filter.ref = f"A1:{get_column_letter(len(basliklar))}1"
        
        # Satırları dondur
        self.ws.freeze_panes = "A2"
    
    def _ozet_istatistikler_olustur(self, sikayetler):
        """Özet istatistikler sayfası"""
        ws = self.wb.create_sheet("Özet İstatistikler")
        
        # Başlık
        ws['A1'] = "ŞİKAYET İSTATİSTİKLERİ"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Toplam sayılar
        ws['A3'] = "Toplam Şikayet:"
        ws['B3'] = len(sikayetler)
        ws['B3'].font = Font(bold=True)
        
        # Durum bazlı sayım
        durum_sayim = {}
        oncelik_sayim = {}
        tur_sayim = {}
        
        for sikayet in sikayetler:
            # Durum
            durum = sikayet[10]
            durum_sayim[durum] = durum_sayim.get(durum, 0) + 1
            
            # Öncelik
            oncelik = sikayet[16] if len(sikayet) > 16 else "Belirtilmemiş"
            oncelik_sayim[oncelik] = oncelik_sayim.get(oncelik, 0) + 1
            
            # Tür
            tur = sikayet[14] if len(sikayet) > 14 else "Belirtilmemiş"
            tur_sayim[tur] = tur_sayim.get(tur, 0) + 1
        
        # Durum tablosu
        row = 5
        ws[f'A{row}'] = "DURUM DAĞILIMI"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for durum, sayi in durum_sayim.items():
            ws[f'A{row}'] = durum
            ws[f'B{row}'] = sayi
            row += 1
        
        # Öncelik tablosu
        row += 2
        ws[f'A{row}'] = "ÖNCELİK DAĞILIMI"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for oncelik, sayi in oncelik_sayim.items():
            ws[f'A{row}'] = oncelik
            ws[f'B{row}'] = sayi
            row += 1
        
        # Tür tablosu
        row += 2
        ws[f'A{row}'] = "ŞİKAYET TÜRLERİ"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # En çok şikayet edilen 5 tür
        sorted_turler = sorted(tur_sayim.items(), key=lambda x: x[1], reverse=True)[:5]
        for tur, sayi in sorted_turler:
            ws[f'A{row}'] = tur
            ws[f'B{row}'] = sayi
            row += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _grafikler_olustur(self, sikayetler):
        """Grafikler sayfası"""
        ws = self.wb.create_sheet("Grafikler")
        
        # Durum dağılımı için veri hazırla
        durum_sayim = {}
        for sikayet in sikayetler:
            durum = sikayet[10]
            durum_sayim[durum] = durum_sayim.get(durum, 0) + 1
        
        # Veriyi sayfaya yaz
        ws['A1'] = "Durum"
        ws['B1'] = "Adet"
        
        row = 2
        for durum, sayi in durum_sayim.items():
            ws[f'A{row}'] = durum
            ws[f'B{row}'] = sayi
            row += 1
        
        # Pasta grafik oluştur
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Durum Dağılımı"
        
        ws.add_chart(pie, "D2")


def test_rapor():
    """Test fonksiyonu"""
    # Örnek veri
    test_sikayetler = [
        (1, "IPT/2024-00001", "Ahmet Yılmaz", "2024-01-15", "İstanbul-Ankara", "ABC123", "555-1234", "Web", "Klima çalışmıyordu", "2024-01-15 10:00", "Yeni", "555-1234", "ahmet@email.com", "34 ABC 123", "Hijyen ve Temizlik", "", "Yüksek"),
        (2, "IPT/2024-00002", "Ayşe Demir", "2024-01-16", "Ankara-İzmir", "DEF456", "555-5678", "Web", "Rötar oldu", "2024-01-16 11:00", "İşlemde", "555-5678", "ayse@email.com", "06 DEF 456", "Rötar / Sefer İptali", "", "Orta"),
    ]
    
    rapor = ExcelRaporlama()
    dosya = rapor.rapor_olustur(test_sikayetler)
    print(f"Rapor oluşturuldu: {dosya}")


if __name__ == "__main__":
    test_rapor()
